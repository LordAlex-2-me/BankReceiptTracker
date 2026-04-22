import openpyxl
from datetime import datetime
import os

import os

# Always save Excel file next to the .exe, wherever it's run from
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'transactions.xlsx')

EXCEL_FILE = 'transactions.xlsx'

def initialize_excel():
    """Create the Excel file with headers if it doesn't exist yet."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Write the header row
        ws.append(['Date', 'Description', 'Type', 'Amount', 'Balance'])

        # Start with a balance of 0 — you can change this later
        ws.append([
            datetime.today().strftime('%Y-%m-%d'),
            'Opening Balance',
            '-',
            0.00,
            0.00
        ])

        wb.save(EXCEL_FILE)
        print(f"Created new file: {EXCEL_FILE}")

def log_transaction(transaction):
    """
    Append a new transaction row and calculate the updated balance.
    transaction = {'type': 'debit', 'amount': 5200.0, 'description': 'Shoprite'}
    """

    # Make sure the file exists first
    initialize_excel()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find the last row that has data
    last_row = ws.max_row

    # Get the previous balance from the last row's Balance column (column 5)
    previous_balance = float(ws.cell(row=last_row, column=5).value or 0)

    # Determine signed amount — debit subtracts, credit adds
    amount = transaction['amount']
    transaction_type = transaction['type']

    # Safety guard — if type couldn't be detected, default to 'unknown'
    if not transaction_type:
        transaction_type = 'unknown'
        signed_amount = 0
    elif transaction_type == 'debit':
        signed_amount = -amount      # money going out
    elif transaction_type == 'credit':
        signed_amount = +amount      # money coming in
    else:
        signed_amount = 0            # unknown, log it but don't affect balance

    # Calculate new balance
    new_balance = previous_balance + signed_amount

    # Get today's date
    date = datetime.today().strftime('%Y-%m-%d')

    # Description fallback if none was found
    description = transaction.get('description') or 'Bank Transaction'

    # Append the new row
    ws.append([date, description, transaction_type.capitalize(), signed_amount, new_balance])

    wb.save(EXCEL_FILE)
    print(f"Logged: {description} | {transaction_type} | {signed_amount} | Balance: {new_balance}")
